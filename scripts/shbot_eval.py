import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import csv

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from langchain_community.chat_models import ChatOllama
from langchain_core.messages import HumanMessage, AIMessage
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import ChatPromptTemplate

import nltk
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
from datasets import load_metric
from sentence_transformers import SentenceTransformer, util
import evaluate
from sklearn.metrics import precision_score, recall_score, f1_score

import datetime
import os

# Initialize language model
llm = ChatOllama(
    base_url="http://127.0.0.1:5000",
    model="sh-llama32",
    keep_alive=-1
)

def evaluate_model(references, generated_responses):
    # Calculate BLEU score
    bleu = evaluate.load("bleu")
    bleu_scores = bleu.compute(predictions=generated_responses, references=references)

    # Calculate ROUGE score
    rouge = evaluate.load('rouge')
    rouge_results = rouge.compute(predictions=generated_responses, references=references)

    return bleu_scores, rouge_results

def extract_json_from_strings(string_list):
    generated_jsons = []
    for i, string in enumerate(string_list):
        string = string.replace("'", '"')
        json_strings = []
        
        # Find the substring before the first '{'
        json_start_index = string.find('{')
        if json_start_index != -1:
            pre_json_string = string[:json_start_index-1]
        else:
            pre_json_string = string
        
        # Extract JSON substrings
        while json_start_index != -1:
            json_end_index = string.find('}', json_start_index)
            if json_end_index == -1:
                break
            json_string = string[json_start_index:json_end_index+1]
            json_strings.append(json_string)
            string = string[json_end_index+1:]
            json_start_index = string.find('{')
        
        if len(json_strings) == 0:
            generated_jsons.append(None)
        elif len(json_strings) == 1:
            try:
                json_object = json.loads(json_strings[0])
                generated_jsons.append(json_object)
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON from string: {json_strings[0]}")
                print(f"Error details: {e}")
                generated_jsons.append(None)
        else:
            # Handle multiple JSONs in the same string
            print(f"Multiple JSONs found in string: {string}")
            error_json = {"error": "response contains multiple JSONs which is invalid"}
            generated_jsons.append(error_json)
        
        # Replace the string in string_list with the substring before the first '{'
        string_list[i] = pre_json_string
    
    return generated_jsons

def normalize_value(value):
    """Normalize the value for comparison."""
    try:
        # Try to convert strings that represent numbers to float
        return float(value)
    except (ValueError, TypeError):
        # If it's not a number or it's already a number, return it as is
        return value

def compare_jsons(generated_json, expected_json):
    """Compare two JSON objects with normalized values."""
    if generated_json is None or expected_json is None:
        return generated_json == expected_json
    
    for key in expected_json:
        if key not in generated_json:
            return False
        # normalize value if the key is "value"
        if key == "value":
            if normalize_value(generated_json[key]) != normalize_value(expected_json[key]):
                return False
        else:
            if generated_json[key] != expected_json[key]:
                return False
    return True

def evaluate_jsons(generated_responses, generated_jsons, expected_json_values):
    correct_count = 0
    total_count = len(generated_responses)
    total_keys = 0
    correct_keys = 0
    json_accuracy_flags = []

    for response, generated_json, expected_json in zip(generated_responses, generated_jsons, expected_json_values):
        if expected_json is not None and isinstance(expected_json, str):
            try:
                expected_json = json.loads(expected_json)
            except json.JSONDecodeError:
                print(f"Error decoding JSON from expected value: {expected_json}")
                json_accuracy_flags.append(False)
                continue
        
        if expected_json is None and generated_json is None:
            correct_count += 1
            json_accuracy_flags.append(True)
            continue

        if expected_json is None:
            if generated_json.get("action") == "none": 
                correct_count += 1
                json_accuracy_flags.append(True)
                print("Expected JSON is None and bot responded with a none action in response:", response)
                continue
            print("Expected JSON is None for response:", response)
            json_accuracy_flags.append(False)
            continue
        
        if generated_json is None:
            print("Generated JSON is None for response:", response)
            json_accuracy_flags.append(False)
            continue

        try:
            keys_correct = compare_jsons(generated_json, expected_json)
            if keys_correct:
                correct_count += 1
                json_accuracy_flags.append(True)
            else:
                json_accuracy_flags.append(False)
            
            for key in expected_json:
                total_keys += 1
                if normalize_value(generated_json.get(key)) == normalize_value(expected_json.get(key)):
                    correct_keys += 1
        except AttributeError:
            print("Error comparing JSON for response:", response)
            json_accuracy_flags.append(False)
    
    accuracy = correct_count / total_count
    key_accuracy = correct_keys / total_keys if total_keys > 0 else 0

    return accuracy, key_accuracy, json_accuracy_flags

def calculate_semantic_similarity(references, generated_responses):
    model = SentenceTransformer('paraphrase-MiniLM-L6-v2')
    embeddings1 = model.encode(references, convert_to_tensor=True)
    embeddings2 = model.encode(generated_responses, convert_to_tensor=True)
    cosine_scores = util.pytorch_cos_sim(embeddings1, embeddings2)
    
    similarities = [cosine_scores[i][i].item() for i in range(len(references))]
    average_similarity = sum(similarities) / len(similarities)
    return similarities, average_similarity

def load_evaluation_dataset_from_json(filename):
    with open(filename, 'r', encoding='utf-8') as jsonfile:
        data = json.load(jsonfile)
    
    def replace_none(value):
        return None if value == "None" else value
    
    inputs = [replace_none(item['Input']) for item in data]
    expected_outputs = [replace_none(item['Expected Output']) for item in data]
    expected_json_values = [replace_none(item['Expected JSON Output']) for item in data]
    
    return inputs, expected_outputs, expected_json_values

def load_evaluation_dataset_from_csv(filename):
    inputs = []
    expected_outputs = []
    expected_json_values = []

    def replace_none(value):
        return None if value == "None" else value

    with open(filename, 'r', encoding='utf-8-sig') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            inputs.append(row['input'])
            expected_outputs.append(row['expected_output'])
            expected_json_values.append(replace_none(row['expected_json']))

    return inputs, expected_outputs, expected_json_values


def process_evaluation_dataset(inputs):
    results = []
    messages_stored = []
    for input_data in inputs:
        messages = parse_messages(input_data)
        messages_stored.append(messages)
        prompt = ChatPromptTemplate.from_messages(messages)
        result = invoke_language_model(llm, prompt)
        results.append(result)
    return results, messages_stored   

def parse_messages(input_data):
    try:
        messages = []
        for message_data in json.loads(input_data):
            if message_data['role'] == 'user':
                messages.append(HumanMessage(content=message_data['content']))
            elif message_data['role'] == 'assistant':
                messages.append(AIMessage(content=message_data['content']))
        return messages
    except json.JSONDecodeError as e:
        print("JSON decode error:", e)
        print("Offending input data:", input_data)
        raise e

def invoke_language_model(llm, prompt):
    chain = prompt | llm | StrOutputParser()
    result = chain.invoke({})
    return result

def calculate_classification_metrics(similarities, json_accuracy_flags, similarity_threshold=0.8):
    y_true = []
    y_pred = []

    for similarity, json_correct in zip(similarities, json_accuracy_flags):
        # True label is positive if JSON is correct
        y_true.append(1 if json_correct else 0)

        # Predicted positive if similarity is above threshold and JSON is correct
        if similarity >= similarity_threshold and json_correct:
            y_pred.append(1)
        else:
            y_pred.append(0)

    # Calculate precision, recall, and F1 score
    precision = precision_score(y_true, y_pred)
    recall = recall_score(y_true, y_pred)
    f1 = f1_score(y_true, y_pred)

    return precision, recall, f1

def write_results_to_excel(inputs, generated_responses, expected_outputs, generated_jsons, expected_json_values, similarities, json_accuracies, average_metrics):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"
    
    # Define header for inputs/outputs
    headers = [
        "inputs", "generated outputs", "expected outputs", 
        "generated jsons", "expected jsons", 
        "semantic similarities", "is_json_correct"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        column_width = len(header) + 5  # Adjust column width to fit the header
        ws.column_dimensions[get_column_letter(col_num)].width = column_width
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")  # Dark grey background for headers
        cell.font = Font(bold=True, color="FFFFFF")  # White font for headers
    
    # Write data rows for inputs/outputs
    for i in range(len(inputs)):
        row = [
            inputs[i], generated_responses[i], expected_outputs[i], 
            json.dumps(generated_jsons[i]), json.dumps(expected_json_values[i]), 
            similarities[i], json_accuracies[i]
        ]
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=i + 2, column=col_num, value=value)
            cell.fill = PatternFill(start_color="dcdcdc", end_color="dcdcdc", fill_type="solid")  # Light grey background for data rows

    # Apply table for inputs/outputs with normal borders
    tab = Table(displayName="EvaluationResults", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Write average metrics on the right in a separate table
    avg_col_start = len(headers) + 3
    avg_headers = ["Metric", "Average Value"]
    
    # Write the average table header
    for col_num, header in enumerate(avg_headers, avg_col_start):
        cell = ws.cell(row=1, column=col_num, value=header)
        column_width = len(header) + 2  # Adjust column width to fit the header
        ws.column_dimensions[get_column_letter(col_num)].width = column_width
    
    # Write the average values
    avg_metrics_items = list(average_metrics.items())
    for row_num, (metric, value) in enumerate(avg_metrics_items, 2):
        ws.cell(row=row_num, column=avg_col_start, value=metric)
        ws.cell(row=row_num, column=avg_col_start + 1, value=value)
    
    # Increase the width of the metric column to 18 (approximately 131 pixels)
    ws.column_dimensions[get_column_letter(avg_col_start)].width = 18

    # Apply table for average metrics with color #327ba8 and no borders
    avg_table_range = f"{get_column_letter(avg_col_start)}1:{get_column_letter(avg_col_start + 1)}{len(avg_metrics_items) + 1}"
    for row in ws[avg_table_range]:
        for cell in row:
            cell.fill = PatternFill(start_color="327ba8", end_color="327ba8", fill_type="solid")
            cell.font = Font(color="FFFFFF")
    
    # Get current timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    # Construct filename with timestamp
    filename = f"evaluation_results_{timestamp}.xlsx"

    try:
        wb.save(filename)
    except PermissionError as e:
        print(f"Permission denied while saving the file: {e}")
        # Append a character to indicate another version of the file
        filename, ext = os.path.splitext(filename)
        filename += "_v2" + ext
        wb.save(filename)
        print(f"File saved as: {filename}")




# MAIN CODE
print("############# LOADING DATASET #############")
#inputs, expected_outputs, expected_json_values = load_evaluation_dataset_from_json('evaluation_dataset.json')
inputs, expected_outputs, expected_json_values = load_evaluation_dataset_from_csv('evaluation_dataset_final.csv')
print(inputs)
print("-------------------------------------------")
print(expected_outputs)
print("-------------------------------------------")
print(expected_json_values)
for x in expected_json_values:
    print(type(x))
print("-------------------------------------------")

# Generate responses from model
generated_responses, messages_stored = process_evaluation_dataset(inputs)
print(generated_responses)
print(messages_stored)

# Extract JSONs from generated responses
print("############# responses and jsons #############")
generated_jsons = extract_json_from_strings(generated_responses)
print(generated_responses)
print(generated_jsons)

# Calculate BLEU and ROUGE metrics
print("############# BLEU/ROUGE #############")
bleus, rouge_results = evaluate_model(expected_outputs, generated_responses)
print(bleus)
print(rouge_results)

# Evaluate semantic similarity
print("############# SEMANTIC SIMILARITY #############")
similarities, average_similarity = calculate_semantic_similarity(expected_outputs, generated_responses)
print(similarities)
print(average_similarity)

# Calculate JSON accuracy
print("############# JSON ACCURACY #############")
json_accuracy, json_key_accuracy, json_accuracy_flags = evaluate_jsons(generated_responses, generated_jsons, expected_json_values)
print(json_accuracy)
print(json_key_accuracy)

# Calculate classification metrics based on semantic similarity and JSON accuracy
print("############# CLASSIFICATION METRICS BASED ON SEMANTIC SIMILARITY AND JSON ACCURACY #############")
precision, recall, f1 = calculate_classification_metrics(similarities, json_accuracy_flags, similarity_threshold=0.65)
print("Precision:", precision)
print("Recall:", recall)
print("F1 Score:", f1)

# Average metrics
average_metrics = {
    "average_bleu": bleus['bleu'],
    "average_rouge1": rouge_results['rouge1'],
    "average_rouge2":rouge_results['rouge2'],
    "average_rougeL": rouge_results['rougeL'],
    "average_similarity": average_similarity,
    "json_accuracy": json_accuracy,
    "json_key_accuracy": json_key_accuracy,
    "precision": precision,
    "recall": recall,
    "f1_score": f1
}


print("############# WRITING RESULTS TO EXCEL #############")
write_results_to_excel(inputs, generated_responses, expected_outputs, generated_jsons, expected_json_values, similarities, json_accuracy_flags, average_metrics)